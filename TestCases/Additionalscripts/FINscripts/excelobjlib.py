import openpyxl

class xl:

    def __init__(self,wb,sh):
        global workbook,sheet
        workbook=openpyxl.load_workbook(wb)
        sheet=workbook[sh]


    def no_of_rows(self):
        return sheet.max_row

    def no_of_cols(self):
        return sheet.max_column

    def get_keys(self):
        lst=[]
        for i in range(1,self.no_of_cols()+1):
            lst.append(sheet.cell(row=1,column=i).value)
        return lst

    def get_data_per_row(self,json_input,rownum,keys):
        #print(json_input)
        for i in range(1,self.no_of_cols()+1):
            #json_input=str(json_input)
            #json_input=json_input.encode('ascii','ignore')
            #print(type(sheet.cell(row=rownum,column=i).value))
            #print(strn)
            strn=sheet.cell(row=rownum,column=i).value
            if(keys[i-1]=='creditParty' or keys[i-1]=='debitParty'):
                #print(keys[i-1])
                json_input[keys[i-1]][0]['value']=strn
                #print(keys[i-1]+":"+json_input[keys[i-1]][0]['value'])
            elif(keys[i-1]=='debitPartyCredentials'):
                #print(keys[i-1])
                json_input[keys[i-1]]['pin']=strn
            else:
                json_input[keys[i-1]]=strn
            #json_input[keys[i-1]].encode('ascii','ignore')
            #print(type(json_input[keys[i-1]]))
        #print(type(json_input))
        #print json_input
        return json_input
